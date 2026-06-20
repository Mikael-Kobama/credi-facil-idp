import os
import json
import io
import boto3
from datetime import datetime
from aws_lambda_powertools import Logger
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = Logger(service="excel-generator")
s3_client = boto3.client("s3", region_name="us-east-1")

def aplicar_estilo_corporativo(ws):
    """Aplica o Design System do CrediFácil na folha do Excel."""
    # 🔥 Força a exibição das linhas de grade nativas do Excel
    ws.views.sheetView[0].showGridLines = True
    
    # 🎨 Paleta de Cores Corporativa
    azul_marinho_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    cinza_claro_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    fonte_cabecalho = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    fonte_comum = Font(name="Arial", size=10, bold=False, color="334155")
    fonte_negrito = Font(name="Arial", size=10, bold=True, color="0F172A")
    
    alinhamento_esquerda = Alignment(horizontal="left", vertical="center")
    alinhamento_centro = Alignment(horizontal="center", vertical="center")
    
    # Bordas finas de contabilidade
    borda_fina = Side(border_style="thin", color="CBD5E1")
    caixa_borda = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)

    # Estiliza o Cabeçalho (Linha 1)
    ws.row_dimensions[1].height = 28
    for cell in ws[1]:
        cell.fill = azul_marinho_fill
        cell.font = fonte_cabecalho
        cell.alignment = alinhamento_centro
        cell.border = caixa_borda

    # Estiliza as Linhas de Dados
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3), start=2):
        ws.row_dimensions[row_idx].height = 20
        # Aplica efeito zebra intercalando cores de fundo
        fill_atual = cinza_claro_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)
        
        for col_idx, cell in enumerate(row, start=1):
            cell.fill = fill_atual
            cell.border = caixa_borda
            if col_idx == 1:
                cell.font = fonte_negrito
                cell.alignment = alinhamento_esquerda
            elif col_idx == 2:
                cell.font = fonte_comum
                cell.alignment = alinhamento_esquerda
            else:
                cell.font = fonte_comum
                cell.alignment = alinhamento_centro

def auto_ajustar_largura_colunas(ws):
    """Calcula dinamicamente a largura ideal das colunas para evitar textos cortados (###)."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 70)

def handler(event, context):
    """Lambda encarregada de ler JSON estruturado do IDP e gerar planilhas executivas .xlsx."""
    try:
        logger.info(f"Iniciando engine de renderização de planilhas para o evento: {json.dumps(event)}")
        
        # Resgata os caminhos de dados mapeados dinamicamente pela nossa esteira
        package_id = event.get("package_id")
        s3_key_json = event.get("s3_key_resultado")
        bucket = event.get("bda_output_bucket") or os.environ.get("BUCKET_SAIDA", "credifacil-docs-saida-dev")
        arquivo_original = event.get("arquivo_original", "documento_analisado.pdf")
        
        if not s3_key_json or not package_id:
            raise ValueError("Propriedades 's3_key_resultado' ou 'package_id' ausentes no payload.")

        # 1. Baixa o JSON estruturado direto do S3 de saída
        s3_response = s3_client.get_object(Bucket=bucket, Key=s3_key_json)
        payload_dados = json.loads(s3_response["Body"].read().decode("utf-8"))
        
        # Captura o nó real de campos extraídos dependendo se é um doc isolado ou consolidado
        campos_extraidos = payload_dados.get("dados_extraidos_do_documento", {})
        if not campos_extraidos and "cliente" in payload_dados:
            # Caso o analista tenha clicado no Excel do relatório consolidado do cliente
            campos_extraidos = event.get("campos_extraidos", {})

        # 2. Inicializa o motor de escrita do OpenPyXL
        wb = Workbook()
        ws = wb.active
        ws.title = "Metadados Extraídos"
        
        # Injeta a estrutura limpa de 3 colunas padrão de auditoria
        ws.append(["Propriedade Analisada", "Valor Identificado pela IA", "Acurácia do Campo"])

        # 3. Varre e transcreve os campos de forma humanizada e limpa
        for chave, bloco_campo in campos_extraidos.items():
            nome_campo = str(chave).replace("_", " ").title()
            
            if isinstance(bloco_campo, dict):
                valor = bloco_campo.get("value")
                # Caso o campo seja um nó/array aninhado complexo
                if valor is None: valor = json.dumps(bloco_campo, ensure_ascii=False)
                
                confianca = bloco_campo.get("confidence", 1.0)
                txt_confianca = f"{float(confianca) * 100:.1f}%" if confianca else "100.0%"
            else:
                valor = bloco_campo if bloco_campo is not None else ""
                txt_confianca = "100.0%"
                
            ws.append([nome_campo, str(valor), txt_confianca])

        # 4. Aplica a camada estética sênior de design
        aplicar_estilo_corporativo(ws)
        auto_ajustar_largura_colunas(ws)

        # 5. Salva o binário físico em memória limpa (BytesIO)
        output_buffer = io.BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)

        # 6. Grava de forma 100% governada dentro da pasta 'planilhas/' no S3
        nome_limpo_arquivo = arquivo_original.replace(".pdf", "").replace(".png", "").replace(".jpg", "")
        s3_target_excel_key = f"results/planilhas/{package_id}/excel_metadados_{nome_limpo_arquivo}.xlsx"
        
        logger.info(f"Salvando planilha legítima .xlsx no storage S3: {s3_target_excel_key}")
        s3_client.put_object(
            Bucket=bucket,
            Key=s3_target_excel_key,
            Body=output_buffer.getvalue(),
            ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        return {
            "status": "SUCCESS",
            "package_id": package_id,
            "excel_s3_key": s3_target_excel_key
        }

    except Exception as e:
        logger.error(f"Falha ao gerar planilha executiva no backend: {str(e)}")
        return {"status": "ERROR", "message": str(e)}