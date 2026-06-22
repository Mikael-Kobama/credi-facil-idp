import os
import json
import io
import boto3
from datetime import datetime
from aws_lambda_powertools import Logger
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = Logger(service="excel-generator")
s3_client = boto3.client("s3", region_name="us-east-1")

def aplicar_estilo_corporativo(ws):
    """Aplica o Design System do CrediFácil na folha do Excel em 2 colunas."""
    ws.views.sheetView[0].showGridLines = True
    
    azul_marinho_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    cinza_claro_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    fonte_cabecalho = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    fonte_comum = Font(name="Arial", size=10, bold=False, color="334155")
    fonte_negrito = Font(name="Arial", size=10, bold=True, color="0F172A")
    
    alinhamento_esquerda = Alignment(horizontal="left", vertical="center")
    alinhamento_centro = Alignment(horizontal="center", vertical="center")
    
    borda_fina = Side(border_style="thin", color="CBD5E1")
    caixa_borda = Border(left=borda_fina, right=borda_fina, top=borda_fina, bottom=borda_fina)

    # Estiliza o Cabeçalho (Linha 1) - Restrito a 2 colunas
    ws.row_dimensions[1].height = 28
    for cell in ws[1]:
        cell.fill = azul_marinho_fill
        cell.font = fonte_cabecalho
        cell.alignment = alinhamento_centro
        cell.border = caixa_borda

    # Estiliza as Linhas de Dados (Apenas colunas 1 e 2)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2), start=2):
        ws.row_dimensions[row_idx].height = 20
        fill_atual = cinza_claro_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)
        
        for col_idx, cell in enumerate(row, start=1):
            cell.fill = fill_atual
            cell.border = caixa_borda
            if col_idx == 1:
                cell.font = fonte_negrito
                cell.alignment = alinhamento_esquerda
            else:
                cell.font = fonte_comum
                cell.alignment = alinhamento_esquerda

def auto_ajustar_largura_colunas(ws):
    """Calcula dinamicamente a largura ideal das colunas."""
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 15), 70)

def handler(event, context):
    try:
        logger.info(f"Iniciando engine de renderização de planilhas para o evento: {json.dumps(event)}")
        
        package_id = event.get("package_id")
        s3_key_json = event.get("s3_key_resultado")
        bucket = event.get("bda_output_bucket") or os.environ.get("BUCKET_SAIDA", "credifacil-docs-saida-dev")
        arquivo_original = event.get("arquivo_original", "documento_analisado.pdf")
        
        if not s3_key_json or not package_id:
            raise ValueError("Propriedades 's3_key_resultado' ou 'package_id' ausentes no payload.")

        s3_response = s3_client.get_object(Bucket=bucket, Key=s3_key_json)
        payload_dados = json.loads(s3_response["Body"].read().decode("utf-8"))
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Metadados Estruturados"
        
        # 🎯 REMOÇÃO CONQUISTADA: Cabeçalho limpo com apenas duas colunas, sem acurácia
        ws.append(["Propriedade Analisada", "Valor Identificado"])

        # 🚀 CASO MESTRE: Se for o JSON consolidado do cliente, monta uma visão executiva sob medida
        if "cliente" in payload_dados and "validacao" in payload_dados:
            cliente = payload_dados["cliente"]
            validacao = payload_dados["validacao"]
            
            ws.append(["Nome Completo do Proponente", cliente.get("nome", "Não Identificado")])
            ws.append(["Documento de Identificação", cliente.get("documento_identificacao", "Não Informado")])
            ws.append(["Score de Crédito Atribuído", f"{cliente.get('score_credito', {}).get('valor', 0)} Pontos"])
            ws.append(["Classificação de Risco", str(cliente.get("classificacao_risco", {}).get("categoria", "INCONCLUSIVO")).upper()])
            ws.append(["Parecer / Justificativa Técnica", cliente.get("classificacao_risco", {}).get("justificativa", "")])
            
            for chk_chave, chk_val in validacao.items():
                nome_chk = str(chk_chave).replace("_", " ").title()
                status_txt = "✅ CONSISTENTE / PRESENTE" if chk_val is True else "❌ DIVERGENTE / AUSENTE" if chk_val is False else "⚪ NÃO AVALIADO"
                ws.append([f"Checklist: {nome_chk}", status_txt])
        
        # CASO PADRÃO: Documentos normais do lote (Holerites, IDs, Cheques)
        else:
            campos_extraidos = payload_dados.get("dados_extraidos_do_documento", {})
            for chave, bloco_campo in campos_extraidos.items():
                nome_campo = str(chave).replace("_", " ").title()
                if isinstance(bloco_campo, dict):
                    valor = bloco_campo.get("value")
                    if valor is None: valor = json.dumps(bloco_campo, ensure_ascii=False)
                else:
                    valor = bloco_campo if bloco_campo is not None else ""
                ws.append([nome_campo, str(valor)])

        aplicar_estilo_corporativo(ws)
        auto_ajustar_largura_colunas(ws)

        output_buffer = io.BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)

        nome_limpo_arquivo = arquivo_original.replace(".pdf", "").replace(".png", "").replace(".jpg", "").replace(".jpeg", "")
        s3_target_excel_key = f"results/planilhas/{package_id}/excel_metadados_{nome_limpo_arquivo}.xlsx"
        
        logger.info(f"Salvando planilha legítima de 2 colunas no S3: {s3_target_excel_key}")
        s3_client.put_object(
            Bucket=bucket, Key=s3_target_excel_key, Body=output_buffer.getvalue(),
            ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        return {"status": "SUCCESS", "package_id": package_id, "excel_s3_key": s3_target_excel_key}

    except Exception as e:
        logger.error(f"Falha ao gerar planilha executiva no backend: {str(e)}")
        raise e